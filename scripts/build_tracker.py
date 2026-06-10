# Builds "Ledger" — a small-business budget & cashflow tracker (.xlsx)
# with live formulas, a 12-month dashboard, and category breakdowns.
# Run: python scripts/build_tracker.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

NAVY="FF1E2761"; ICE="FFCADCFC"; LITE="FFF4F7FE"; INK="FF2A2E3A"; MUT="FF6B7280"
WHITE="FFFFFFFF"; GREEN="FF1F7A4D"; RED="FFB4232B"
MONTHS=["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
thin=Side(style="thin", color="FFD8DEEC")
box=Border(left=thin,right=thin,top=thin,bottom=thin)
H=Font(name="Calibri",bold=True,color=WHITE,size=11)
TITLE=Font(name="Georgia",bold=True,color=NAVY,size=20)
SUB=Font(name="Calibri",color=MUT,size=10)
LBL=Font(name="Calibri",color=INK,size=11)
BOLD=Font(name="Calibri",bold=True,color=NAVY,size=11)
navy_fill=PatternFill("solid",fgColor=NAVY)
ice_fill=PatternFill("solid",fgColor=ICE)
lite_fill=PatternFill("solid",fgColor=LITE)
ctr=Alignment(horizontal="center",vertical="center")
rt=Alignment(horizontal="right")
MONEY='#,##0;[Red]-#,##0'

wb=openpyxl.Workbook()

# ---------- Sheet 1: Dashboard ----------
d=wb.active; d.title="Dashboard"; d.sheet_view.showGridLines=False
d.column_dimensions["A"].width=3
for c in "BCDE": d.column_dimensions[c].width=20
d["B2"]="LEDGER"; d["B2"].font=Font(name="Calibri",bold=True,color=NAVY,size=12)
d["B3"]="Budget & cashflow tracker"; d["B3"].font=TITLE
d["B4"]="Fill in Income and Expenses — every figure below updates itself."; d["B4"].font=SUB

# KPI cards
kpis=[("B","Total income","=SUM(Income!B3:M30)"),
      ("C","Total expenses","=SUM(Expenses!B3:M40)"),
      ("D","Net profit","=Dashboard!B7-Dashboard!C7"),
      ("E","Profit margin","=IF(B7=0,0,D7/B7)")]
for col,label,formula in kpis:
    d[f"{col}6"]=label; d[f"{col}6"].font=Font(name="Calibri",bold=True,color=WHITE,size=10); d[f"{col}6"].fill=navy_fill; d[f"{col}6"].alignment=ctr
    cell=d[f"{col}7"]; cell.value=formula; cell.font=Font(name="Georgia",bold=True,color=NAVY,size=16); cell.alignment=ctr; cell.fill=lite_fill
    cell.number_format='0.0%' if col=="E" else MONEY
    d[f"{col}6"].border=box; d[f"{col}7"].border=box
d.row_dimensions[7].height=34

# Monthly net table (feeds chart)
d["B10"]="Month"; d["C10"]="Income"; d["D10"]="Expenses"; d["E10"]="Net"
for c in "BCDE":
    d[f"{c}10"].font=H; d[f"{c}10"].fill=navy_fill; d[f"{c}10"].alignment=ctr; d[f"{c}10"].border=box
for i,m in enumerate(MONTHS):
    r=11+i; col=get_column_letter(2+i)  # B..M in source sheets
    d[f"B{r}"]=m; d[f"B{r}"].font=LBL; d[f"B{r}"].border=box
    d[f"C{r}"]=f"=SUM(Income!{col}3:{col}30)"
    d[f"D{r}"]=f"=SUM(Expenses!{col}3:{col}40)"
    d[f"E{r}"]=f"=C{r}-D{r}"
    for c in "CDE":
        d[f"{c}{r}"].number_format=MONEY; d[f"{c}{r}"].font=LBL; d[f"{c}{r}"].border=box
        if i%2: d[f"{c}{r}"].fill=lite_fill
    if i%2: d[f"B{r}"].fill=lite_fill

chart=BarChart(); chart.type="col"; chart.title="Income vs Expenses by month"; chart.height=7; chart.width=16
data=Reference(d,min_col=3,max_col=4,min_row=10,max_row=22)
cats=Reference(d,min_col=2,min_row=11,max_row=22)
chart.add_data(data,titles_from_data=True); chart.set_categories(cats)
d.add_chart(chart,"G6")
line=LineChart(); line.title="Net profit trend"; line.height=7; line.width=16
ld=Reference(d,min_col=5,max_col=5,min_row=10,max_row=22)
line.add_data(ld,titles_from_data=True); line.set_categories(cats)
d.add_chart(line,"G21")

# ---------- Sheet 2: Income ----------
inc=wb.create_sheet("Income"); inc.sheet_view.showGridLines=False
inc.column_dimensions["A"].width=26
for i in range(12): inc.column_dimensions[get_column_letter(2+i)].width=10
inc.column_dimensions["N"].width=12
inc["A1"]="INCOME — fill in each source, one row each"; inc["A1"].font=TITLE
hdr=["Source"]+MONTHS+["Total"]
for j,h in enumerate(hdr):
    c=inc.cell(row=2,column=1+j,value=h); c.font=H; c.fill=navy_fill; c.alignment=ctr; c.border=box
sample=["Product sales","Services","Retainers","Other"]
for r in range(3,31):
    inc.cell(row=r,column=1,value=sample[r-3] if r-3<len(sample) else None).font=LBL
    inc.cell(row=r,column=1).border=box
    for cc in range(2,14):
        cell=inc.cell(row=r,column=cc); cell.number_format=MONEY; cell.border=box
        if r%2==0: cell.fill=lite_fill
    t=inc.cell(row=r,column=14,value=f"=SUM(B{r}:M{r})"); t.number_format=MONEY; t.font=BOLD; t.border=box; t.fill=ice_fill
    if r%2==0: inc.cell(row=r,column=1).fill=lite_fill

# ---------- Sheet 3: Expenses ----------
exp=wb.create_sheet("Expenses"); exp.sheet_view.showGridLines=False
exp.column_dimensions["A"].width=26
for i in range(12): exp.column_dimensions[get_column_letter(2+i)].width=10
exp.column_dimensions["N"].width=12
exp["A1"]="EXPENSES — fill in each cost, one row each"; exp["A1"].font=TITLE
for j,h in enumerate(hdr):
    c=exp.cell(row=2,column=1+j,value=h); c.font=H; c.fill=navy_fill; c.alignment=ctr; c.border=box
ex_sample=["Salaries","Rent","Software & tools","Marketing","Contractors","Taxes","Other"]
for r in range(3,41):
    exp.cell(row=r,column=1,value=ex_sample[r-3] if r-3<len(ex_sample) else None).font=LBL
    exp.cell(row=r,column=1).border=box
    for cc in range(2,14):
        cell=exp.cell(row=r,column=cc); cell.number_format=MONEY; cell.border=box
        if r%2==0: cell.fill=lite_fill
    t=exp.cell(row=r,column=14,value=f"=SUM(B{r}:M{r})"); t.number_format=MONEY; t.font=BOLD; t.border=box; t.fill=ice_fill
    if r%2==0: exp.cell(row=r,column=1).fill=lite_fill

import os
os.makedirs("templates/tracker",exist_ok=True)
wb.save("templates/tracker/ledger-tracker.xlsx")
print("tracker OK")
